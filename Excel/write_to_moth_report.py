import openpyxl
import os
from Read_PR_report import result
target_path_report = 'C:/Users/borisov/Desktop/TEST.xlsx'
path_normalized_report = os.path.normpath(target_path_report)

report = openpyxl.load_workbook(path_normalized_report)
print(report.sheetnames)
sheet = report.active
print(sheet.title)
sheet.title = 'oufigd'
print(report.sheetnames)
sheet['a4'] = result


report.save('TEST_copy1.xlsx')
