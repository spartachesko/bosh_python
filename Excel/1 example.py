import openpyxl
import os

# target_path = 'C:/Users/borisov/PycharmProjects/bosh_python/Excel/automate_online-materials/censuspopdata.xlsx'
target_path = 'Q:/COMMON DOCS/1st PRODUCTION SECTION/0.Daily monitoring (Ежедневный контроль)/Sewing 2021/' \
              'Март  2021.xlsx'
path_normalized = os.path.normpath(target_path)
wb = openpyxl.load_workbook(path_normalized)
list_of_sheets = wb.sheetnames




print(list_of_sheets)
sheet = wb['Total production']
print(sheet.title)
production = sheet['F6'].value
print(production)

print('Введите месяц, за который хотите узнать количество произведенных сетов.')
print('Ввод осуществлять в формате: мар.21')
month = str(input())


def choose_date(month=month):
    result = None
    if month == 'январь':
        result = sheet.cell(row=6, column=4).value
        return print(result, ' сетов')

    elif month == 'февраль':
        result = sheet.cell(row=6, column=5).value
        return print(result, ' сетов')
    elif month == 'март':
        result = sheet.cell(row=6, column=6).value
        return print(result, ' сетов')
    elif month == 'апрель':
        result = sheet.cell(row=6, column=7).value
        return print(result, ' сетов')
    elif month == 'май':
        result = sheet.cell(row=6, column=8).value
        return print(result, ' сетов')
    elif month == 'июнь':
        result = sheet.cell(row=6, column=9).value
        return print(result, ' сетов')
    elif month == 'июль':
        result = sheet.cell(row=6, column=10).value
        return print(result, ' сетов')
    elif month == 'август':
        result = sheet.cell(row=6, column=11).value
        return print(result, ' сетов')
    elif month == 'сентябрь':
        result = sheet.cell(row=6, column=12).value
        return print(result, ' сетов')
    elif month == 'октябрь':
        result = sheet.cell(row=6, column=13).value
        return print(result, ' сетов')
    elif month == 'ноябрь':
        result = sheet.cell(row=6, column=14).value
        return print(result, ' сетов')

    elif month == 'декабрь':
        result = sheet.cell(row=6, column=15).value
        return print(result, ' сетов')
    else:
        return print('некорректный ввод ')


# def choose_date(month=month):
#     result = None
#     all_dates = tuple(sheet['D5':'AY5'])
#     for cell_of_dates in all_dates:
#         if month == cell_of_dates:
#             coord_date = cell_of_dates.coordinate
#             return print('Количество сетов: ', coord_date)
#         else:
#             return print('Нет такого месяца')

choose_date(month=month)
sheet_report = report.sheetnames