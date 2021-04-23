import openpyxl
import os

target_path_report = 'C:/Users/borisov/Desktop/TEST.xlsx'
path_normalized_report = os.path.normpath(target_path_report)
# report = openpyxl.load_workbook(path_normalized_report)
# list_of_sheets_report = report.sheetnames

# sheet_report = report['SEW_Plan-Report(mar) (2)']
# print(sheet_report)
# sheet_report['C16'] = '123456'
# print(sheet_report['C16'].value)

report = openpyxl.Workbook(path_normalized_report)
ws1 = report.create_sheet("Kekukek")
report.save('C:/Users/borisov/Desktop/TEST_copy.xlsx')
